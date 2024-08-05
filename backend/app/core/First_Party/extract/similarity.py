from gensim.models import KeyedVectors
from sklearn.metrics.pairwise import cosine_similarity
import jieba
import numpy as np
import gensim.downloader
import json
import openpyxl
import os
import re

with open('C:\\Users\\air\\Desktop\\大三下\\信安\\ZZ施工队-面向第三方SDK的隐私透明度提升工具\\PrivacyComplyGuard\\First_Party\\E2C.json', 'r', encoding='utf-8') as f:
    word_mapping = json.load(f)

model = KeyedVectors.load_word2vec_format(
    r"C:\Users\air\Desktop\大三下\信安\ZZ施工队-面向第三方SDK的隐私透明度提升工具\PrivacyComplyGuard\models\baike_26g_news_13g_novel_229g.bin",
    binary=True
)
# google_news_vectors = gensim.downloader.load('word2vec-google-news-300')

flag = False

def store_categoryjson():
    # 读取 Excel 文件
    workbook = openpyxl.load_workbook('层次分类.xlsx')
    sheet = workbook.active

    # 创建字典,存储键值对
    data = {}
    for row in range(2, sheet.max_row + 1):
        key = str(sheet.cell(row=row, column=1).value)
        value = str(sheet.cell(row=row, column=2).value)
        data[key] = value

    # 将字典写入 JSON 文件
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_file = os.path.join(current_dir, '分类.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    print("数据已成功写入 JSON 文件.")

def translate_word(word):
    global word_mapping
    new_word_mapping = {k.lower(): v for k, v in word_mapping.items()}  # 键都换成小写
    # key = [k.lower() for k in word_mapping.keys()]
    # print("key=", key)
    if word.lower() in new_word_mapping.keys():
        w = new_word_mapping[word.lower()]
        return w
    else:
        return None

def get_word_vectors(word, model):
    """获取单个词的词向量"""
    global word_mapping
    if word in model:
        return model[word]
    else:
        tran_word = translate_word(word)
        # print("tran_word:", tran_word)
        if tran_word is None:
            return None

        if tran_word in model:
            return model[tran_word]
        else:
            vector = tokenize(tran_word, model)

            if (vector is not None) and (len(vector) == 0):
                return None
            elif(len(vector) == 128):
                # 说明是向量
                return vector
            else:
                vectors = []
                for word in vector:
                    word_vector = get_word_vectors(word, model)
                    if word_vector is not None:
                        vectors.append(word_vector)

                try:
                    re_vector = np.mean(vectors, axis=0)
                    # re_vector = np.sum(vectors, axis=0)
                    return re_vector
                except:
                    print("vectors=", vectors)
                    return None

def tokenize(word, model):
    """递归分词直到找到词向量"""
    global flag
    word_vector = get_word_vectors(word, model)
    if word_vector is not None:
        return [word_vector]
    else:
        # 如果当前词没有词向量，继续分词
        sub_words = list(jieba.cut(word))
        return [sub_word for sub_word in sub_words if get_word_vectors(sub_word, model) is not None]

def get_phrase_vectors(phrase, model):
    """获取词组的词向量，如果词组不存在，则尝试分词并递归查找词向量"""
    phrase_vector = get_word_vectors(phrase, model)
    if phrase_vector is not None:
        return phrase_vector

    # 分词并递归查找词向量
    words = list(jieba.cut(phrase))
    phrase_vectors = []
    for word in words:
        vector = tokenize(word, model)
        if (vector is not None) and (len(vector) > 0):
            phrase_vectors.append(vector)

    # 如果没有找到任何词向量，返回None
    if len(phrase_vectors)==0:
        return None

    # 计算词向量的平均值
    try:
        phrase_vector = np.mean(phrase_vectors, axis=0)
        # phrase_vector = np.sum(phrase_vectors, axis=0)
        return phrase_vector
    except:
        print("phrase_vector=", phrase_vector)
        return None

# 示例使用
def get_phrase_similarity(phrase1, phrase2):

    phrase_vector1 = get_phrase_vectors(phrase1, model)
    phrase_vector2 = get_phrase_vectors(phrase2, model)

    if (phrase_vector1 is None) or (phrase_vector2 is None):
        print("其中有向量是None")
        if phrase_vector1 is None:
            print(phrase1, ":", phrase_vector1)
        if phrase_vector2 is None:
            print(phrase2, ":", phrase_vector2)
        return 0
    if(len(phrase_vector1)==1):
        phrase_vector1 = phrase_vector1[0]
    if(len(phrase_vector2)==1):
        phrase_vector2 = phrase_vector2[0]

    # print(f"{phrase1}':{phrase_vector1}")
    # print(f"{phrase2}':{phrase_vector2}")

    try:
        similarity = cosine_similarity(phrase_vector1.reshape(1, -1), phrase_vector2.reshape(1, -1))[0][0]
        print(f"Cosine similarity between {phrase1} and {phrase2}:{similarity}")
        return similarity
    except:
        print(f"没有对应词向量")
        return 0

def get_all_category():
    '''
    :return: 那个Excel里所有检测出来的缺失声明的类别
    '''
    # category = ["传感器信息", "短信信息", "房产信息", "敏感个人信息", "设备信息", "应用日志信息", "图片信息"]
    workbook = openpyxl.load_workbook('category.xlsx')
    sheet = workbook.active

    # 获取第 6 列数据
    column_data = []
    for row in range(2, sheet.max_row + 1):
    # for row in range(26, 28):
        cell_value = sheet.cell(row=row, column=6).value
        if cell_value:
            cell_value = str(cell_value)
            # 使用正则表达式进行分割
            pattern = r"(?<![(])\、(?![)])"
            new_list = re.split(pattern, cell_value)
            column_data.append(new_list)

    # 获取每个缺失声明的对应类别
    all_category = []

    # 读取已经有的检测结果类别
    with open("所有检测数据项的类别.json", "r", encoding="gbk") as f:
        word_category_dict = json.load(f)

    # word_category_dict = {}
    with open("类别统计.json", "r", encoding="gbk") as f:
        category_count = json.load(f)
    # category_count = {}

    for item in column_data:  # 每个单元格里的列表
        # print(item)
        for word in item:  # 每个单元格里的列表的每个词
            word_category = ''
            if word in word_category_dict:   # 和之前分类过的词重复了
                word_category = word_category_dict[word]

            else:
                word_category = get_category(word, keys, values)
                # all_category.append(word_category)
                word_category_dict[word] = word_category

            # 统计每个类别出现的次数
            if word_category in category_count:
                category_count[word_category] += 1
            else:
                category_count[word_category] = 1

    # 保存单词和类别的对应关系
    with open("所有检测数据项的类别.json", "w") as f:
        json.dump(word_category_dict, f, ensure_ascii=False, indent=4)

    # 保存类别统计结果
    with open("类别统计.json", "w") as f:
        json.dump(category_count, f, ensure_ascii=False, indent=4)

def get_category(word, all_current_item, values):
    '''
    :param word: 新的数据项
    :param all_current_item: 现在有了的数据项（已经分类了的）
    :return: 类别
    '''

    # 调试某个词
    # all_current_item = ["IP地址"]

    # 获取某个词组的大类
    max_similarity = 0
    max_index = 0
    category = values[0]
    record_index = 0  #记录出错的词
    try:
        for index, item in enumerate(all_current_item):
            record_index = index
            s = get_phrase_similarity(word, item)
            if s == 0:
                print(f"新词 {word} 和 {all_current_item[record_index]} 的比较出错了")
                category = "设备信息"
                print(word, " 的类别是：", category)
                print("-" * 20)
                return category    # all_current_item的词都不会有问题了，所以如果出错的话只可能是因为新词是一个英文单词。

            if s > max_similarity:
                max_similarity = s
                max_index = index

        print("和 ", word, " 最接近的词是：", all_current_item[max_index])
        print("相似度是：", max_similarity)
        category = values[max_index]
        print("类别是：", category)
        print("-"*20)
        return category

    except:
        category = "设备信息"
        print(word, " 的类别是：", category)
        print("-" * 20)
        return category


# 度量两个词
phrase1 = "GPS信息"
phrase2 = "位置信息"
get_phrase_similarity(phrase1, phrase2)

# with open('分类.json', 'r', encoding='utf-8') as f:
#     data = json.load(f)
# keys = list(data.keys())
# values = list(data.values())

# 找单个词的类别
# c = get_category(phrase1, keys, values)
# print("类别是：", c)
# store_categoryjson()

# phrase_vector1 = get_phrase_vectors(phrase1, model)
# print("phrase_vector1：", phrase_vector1)

# 找所有词的类别
get_all_category()